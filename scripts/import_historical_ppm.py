#!/usr/bin/env python3
"""
Download and import all historical PPM fund NAV data from Pensionsmyndigheten.
Data covers 2000–2026 in three formats: txt, xls, xlsx.

Usage:
    python3 scripts/import_historical_ppm.py [--db path/to/dashboard.db] [--cache-dir /tmp/ppm_dl]
"""
import argparse
import datetime
import logging
import os
import sys
from pathlib import Path

import requests

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

BASE_URL = (
    "https://www.pensionsmyndigheten.se/content/dam/pensionsmyndigheten/"
    "blanketter---broschyrer---faktablad/statistik/premiepension/fonddata/"
    "historiska-fondkurser/"
)

# Exact filenames as they appear on the server (URL-encoded spaces etc kept as-is)
FILES = [
    (2026, "Fondandelskurser 2026.xlsx",  "xlsx"),
    (2025, "Fondandelskurser 2025.xlsx",  "xlsx"),
    (2024, "Fondandelskurser 2024.xlsx",  "xlsx"),
    (2023, "Fondandelskurser 2023.xlsx",  "xlsx"),
    (2022, "Fondandelskurser 2022.xlsx",  "xlsx"),
    (2021, "Fondandelskurser 2021.xlsx",  "xlsx"),
    (2020, "Fondandelskurser 2020.xlsx",  "xlsx"),
    (2019, "Fondandelskurser 2019.xlsx",  "xlsx"),
    (2018, "Fondandelskurser  2018.xlsx", "xlsx"),  # extra space
    (2017, "Fondandelskurser  2017.xlsx", "xlsx"),  # extra space
    (2016, "Fondandelskurser  2016 .xlsx","xlsx"),  # extra spaces
    (2015, "Fondandelskurser 2015.xlsx",  "xlsx"),
    (2014, "Fondandelskurser 2014.xlsx",  "xlsx"),
    (2013, "Fondandelskurser 2013.xls",   "xls"),
    (2012, "Fondandelskurser 2012.xls",   "xls"),
    (2011, "Fondandelskurser 2011.xls",   "xls"),
    (2010, "fondandelskurser 2010.xls",   "xls"),  # lowercase
    (2009, "Fondandelskurser 2009.xls",   "xls"),
    (2008, "Fondandelskurser 2008.xls",   "xls"),
    (2007, "Fondandelskurser 2007.xls",   "xls"),
    (2006, "Fondandelskurser 2006.xls",   "xls"),
    (2005, "Fondkurser 2005.txt",          "txt"),
    (2004, "Fondkurser 2004.txt",          "txt"),
    (2003, "Fondkurser 2003.txt",          "txt"),
    (2002, "Fondkurser 2002.txt",          "txt"),
    (2001, "Fondkurser 2001.txt",          "txt"),
    (2000, "Fondkurser 2000.txt",          "txt"),
]


# ── Downloaders ────────────────────────────────────────────────────────────────

def download_file(filename: str, cache_dir: Path) -> Path:
    dest = cache_dir / filename
    if dest.exists() and dest.stat().st_size > 10_000:
        log.info("  cached: %s", filename)
        return dest
    url = BASE_URL + requests.utils.quote(filename)
    log.info("  downloading %s …", filename)
    r = requests.get(url, timeout=120, stream=True)
    r.raise_for_status()
    with open(dest, "wb") as f:
        for chunk in r.iter_content(chunk_size=1 << 20):
            f.write(chunk)
    log.info("    → %.1f MB", dest.stat().st_size / 1e6)
    return dest


# ── Parsers ────────────────────────────────────────────────────────────────────

def _safe_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def parse_txt(path: Path) -> list[tuple]:
    """Tab-separated, cp1252. Cols: 0=FONDNUMMER, 4=DATUM, 9=FONDKURS_SEK_KÖP"""
    rows = []
    with open(path, encoding="cp1252", errors="replace") as f:
        next(f)  # skip header
        for line in f:
            parts = line.rstrip("\n").split("\t")
            if len(parts) < 10:
                continue
            ppm   = parts[0].strip()
            date  = parts[4].strip()
            nav   = _safe_float(parts[9])
            if ppm and date and nav and nav > 0:
                rows.append((ppm, date, nav))
    return rows


def parse_xls(path: Path) -> list[tuple]:
    """xlrd. Auto-detects column layout from header row."""
    import xlrd
    MIN_SERIAL = 36526  # Excel serial for 2000-01-01
    wb = xlrd.open_workbook(str(path))
    rows = []
    for sheet in wb.sheets():
        if sheet.nrows < 2:
            continue
        header = [str(sheet.cell(0, c).value).strip().upper() for c in range(sheet.ncols)]
        # 2006: FONDKURS_DATUM in col 8, FONDKURS_SEK_BUY in col 10
        if "FONDKURS_DATUM" in header:
            col_date = header.index("FONDKURS_DATUM")
            col_nav  = next((i for i, h in enumerate(header) if "SEK" in h and "BUY" in h), 10)
        else:
            col_date = 3
            col_nav  = next((i for i, h in enumerate(header) if "SEK" in h and "KÖP" in h.upper()), 8)
        for ri in range(1, sheet.nrows):
            row = sheet.row_values(ri)
            if len(row) <= max(col_date, col_nav):
                continue
            ppm_raw = row[1]
            if not ppm_raw:
                continue
            ppm = str(int(float(ppm_raw))) if isinstance(ppm_raw, float) else str(ppm_raw).strip()
            date_raw = row[col_date]
            if not date_raw or (isinstance(date_raw, float) and date_raw < MIN_SERIAL):
                continue
            try:
                dt = xlrd.xldate_as_datetime(date_raw, wb.datemode)
                date = dt.strftime("%Y-%m-%d")
            except Exception:
                continue
            nav = _safe_float(row[col_nav])
            if ppm and nav and nav > 0:
                rows.append((ppm, date, nav))
    return rows


def parse_xlsx(path: Path) -> list[tuple]:
    """openpyxl. Cols: 1=Fondnummer, 3=Datum, 5=Valutakurs köp, 7=Fondkurs köp → NAV_SEK = 7*5"""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets:
        for ri, row in enumerate(sheet.iter_rows(values_only=True)):
            if ri == 0:
                continue  # header
            if not row or row[1] is None:
                continue
            ppm = str(row[1]).strip()
            date_raw = row[3]
            if isinstance(date_raw, datetime.datetime):
                date = date_raw.strftime("%Y-%m-%d")
            elif isinstance(date_raw, str):
                date = date_raw[:10]
            else:
                continue
            fx    = _safe_float(row[5])  # Valutakurs köp
            kurs  = _safe_float(row[7])  # Fondkurs köp
            if not fx or not kurs:
                continue
            nav = kurs * fx
            if ppm and nav > 0 and date >= "2000-01-01":
                rows.append((ppm, date, nav))
    wb.close()
    return rows


# ── DB upsert ──────────────────────────────────────────────────────────────────

def upsert_rows(rows: list[tuple], db_path: Path) -> int:
    import sqlite3
    con = sqlite3.connect(db_path)
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ppm_nav (
            ppm_number TEXT,
            date       TEXT,
            nav_sek    REAL,
            PRIMARY KEY (ppm_number, date)
        )
    """)
    BATCH = 50_000
    inserted = 0
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        cur.executemany(
            "INSERT OR REPLACE INTO ppm_nav(ppm_number, date, nav_sek) VALUES (?,?,?)",
            batch,
        )
        inserted += len(batch)
    con.commit()
    con.close()
    return inserted


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import historical PPM NAV data")
    parser.add_argument("--db", default=None, help="Path to dashboard.db")
    parser.add_argument("--cache-dir", default="/tmp/ppm_dl", help="Directory to cache downloaded files")
    parser.add_argument("--years", nargs="+", type=int, help="Only import specific years")
    args = parser.parse_args()

    # Resolve DB path
    if args.db:
        db_path = Path(args.db)
    else:
        repo_root = Path(__file__).parent.parent
        data_dir  = Path(os.getenv("DATA_DIR", str(repo_root)))
        db_path   = data_dir / "dashboard.db"

    if not db_path.exists():
        log.error("DB not found: %s", db_path)
        sys.exit(1)

    cache_dir = Path(args.cache_dir)
    cache_dir.mkdir(parents=True, exist_ok=True)

    files = FILES
    if args.years:
        files = [(y, fn, fmt) for y, fn, fmt in FILES if y in args.years]

    total = 0
    for year, filename, fmt in files:
        log.info("── %d (%s) ──────────────────────", year, fmt)
        try:
            path = download_file(filename, cache_dir)
        except Exception as e:
            log.error("  download failed: %s", e)
            continue

        try:
            if fmt == "txt":
                rows = parse_txt(path)
            elif fmt == "xls":
                rows = parse_xls(path)
            else:
                rows = parse_xlsx(path)
        except Exception as e:
            log.error("  parse failed: %s", e)
            continue

        log.info("  parsed %d rows", len(rows))

        try:
            n = upsert_rows(rows, db_path)
            total += n
            log.info("  upserted %d rows → DB", n)
        except Exception as e:
            log.error("  DB insert failed: %s", e)

    log.info("══ Done — %d rows total ══", total)


if __name__ == "__main__":
    main()
